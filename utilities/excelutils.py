"""
@package utilities

ExcelUtils class implementation
All excel utilities should be implemented in this class

"""
import openpyxl

class ExcelUtils(object):

    def getRowCount(self,file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.max_row

    def getColumnCount(self,file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.max_column

    def readDataCell(self,file,sheetName,rowNum,colNum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.cell(row=rowNum,column=colNum).value

    def readAllCellData(self,file,sheetName,firstName,secondName):
        list = []
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        maxRows = sheet.max_row
        maxColumns = sheet.max_column
        for row in range(2,maxRows+1):
            firstName=sheet.cell(row,1).value
            secondName=sheet.cell(row,2).value
            tuple = (firstName,secondName)
            list.append(tuple)





    def writeData(self,file,sheetName,rowNum,colNum,data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        sheet.cell(row=rowNum,column=colNum).value = data
        workbook.save(file)
